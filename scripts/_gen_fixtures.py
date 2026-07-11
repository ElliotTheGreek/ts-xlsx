"""Generate the ts-xlsx test fixture corpus with openpyxl.

These are byte-round-trip oracles: ts-xlsx must parse->serialize each one back
byte-identically. openpyxl-authored files (including chart/image workbooks)
exercise the fidelity backbone. Genuine Excel-authored files (pivot tables,
vendor XML) can be dropped into tests/assets/ later to strengthen the corpus.
"""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.chart import BarChart, Reference
from openpyxl.drawing.image import Image as XLImage
from openpyxl.comments import Comment
import datetime

HERE = os.path.dirname(os.path.abspath(__file__))
ASSETS = os.path.join(HERE, "..", "tests", "assets")
os.makedirs(ASSETS, exist_ok=True)


def p(name):
    return os.path.join(ASSETS, name)


# 1. basic.xlsx — multi-sheet, values, merges, widths, one styled cell
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Data"
ws["A1"] = "Name"
ws["B1"] = "Score"
ws["A2"] = "Alice"
ws["B2"] = 42
ws["A3"] = "Bob"
ws["B3"] = 17.5
ws["C1"] = True
ws["A1"].font = Font(bold=True, color="CC0000", size=14)
ws["B1"].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
ws["B2"].number_format = "#,##0.00"
ws["A1"].alignment = Alignment(horizontal="center")
ws["A1"].border = Border(bottom=Side(style="thin"))
ws.merge_cells("A5:C5")
ws["A5"] = "merged title"
ws.column_dimensions["A"].width = 20
ws.row_dimensions[1].height = 22
ws.freeze_panes = "A2"
ws2 = wb.create_sheet("Notes")
ws2["A1"] = "second sheet"
wb.create_sheet("Hidden").sheet_state = "hidden"
wb.save(p("basic.xlsx"))

# 2. formulas.xlsx — stored formulas
wb = openpyxl.Workbook()
ws = wb.active
for i in range(1, 6):
    ws.cell(row=i, column=1, value=i)
ws["B1"] = "=SUM(A1:A5)"
ws["B2"] = "=A1*2"
ws["C1"] = "=IF(B1>10,\"big\",\"small\")"
wb.save(p("formulas.xlsx"))

# 3. types.xlsx — dates, bools, shared strings (repeated strings)
wb = openpyxl.Workbook()
ws = wb.active
ws["A1"] = datetime.datetime(2026, 7, 11, 9, 30, 0)
ws["A2"] = datetime.date(2026, 1, 1)
ws["A3"] = True
ws["A4"] = False
ws["A5"] = "repeat"
ws["A6"] = "repeat"
ws["A7"] = "repeat"
ws["A8"] = 3.14159
ws["A9"] = "unicode: café ☕ 数"
wb.save(p("types.xlsx"))

# 4. chart.xlsx — a bar chart (openpyxl drops charts on its OWN reload of files
#    it didn't author; ts-xlsx must preserve this one byte-for-byte untouched)
wb = openpyxl.Workbook()
ws = wb.active
ws.append(["Month", "Sales"])
for m, s in [("Jan", 10), ("Feb", 40), ("Mar", 25), ("Apr", 60)]:
    ws.append([m, s])
chart = BarChart()
chart.title = "Sales"
data = Reference(ws, min_col=2, min_row=1, max_row=5)
cats = Reference(ws, min_col=1, min_row=2, max_row=5)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
ws.add_chart(chart, "D2")
wb.save(p("chart.xlsx"))

# 5. image.xlsx — embedded PNG (openpyxl drops images on reload; we must not)
png_1x1 = bytes.fromhex(
    "89504e470d0a1a0a0000000d49484452000000010000000108060000001f15c4"
    "890000000d49444154789c6360000002000100" "05fe02fea7e3b3c00000000049454e44ae426082"
)
img_path = p("_1x1.png")
with open(img_path, "wb") as f:
    f.write(png_1x1)
wb = openpyxl.Workbook()
ws = wb.active
ws["A1"] = "has image"
ws.add_image(XLImage(img_path), "B2")
wb.save(p("image.xlsx"))
os.remove(img_path)

# 6. comments.xlsx — cell comments + data validation-ish styling
wb = openpyxl.Workbook()
ws = wb.active
ws["A1"] = "commented"
ws["A1"].comment = Comment("a note", "author")
wb.save(p("comments.xlsx"))

print("fixtures written to", os.path.normpath(ASSETS))
for f in sorted(os.listdir(ASSETS)):
    if f.endswith(".xlsx"):
        print("  ", f, os.path.getsize(os.path.join(ASSETS, f)), "bytes")
